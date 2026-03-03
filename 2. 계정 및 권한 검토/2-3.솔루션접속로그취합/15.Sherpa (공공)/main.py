import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from config import TARGET_FILE, SOURCE_FOLDER, TARGET_COLUMNS, COLUMN_MAPPING
from config import COLOR_RED, COLOR_BLUE, FONT_SIZE, ALIGN_HORIZONTAL, ALIGN_VERTICAL
from config import DAYS_STOP, DAYS_DELETE, TARGET_SHEET, SOURCE_SHEET

def debug_log(msg):
    print(f"[DEBUG] {msg}")

def ensure_worksheet(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        debug_log(f"Worksheet '{sheet_name}' exists.")
        return workbook[sheet_name]
    debug_log(f"Worksheet '{sheet_name}' not found. Creating...")
    return workbook.create_sheet(sheet_name)

def add_columns(sheet, columns):
    debug_log(f"Adding columns: {columns}")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(size=FONT_SIZE)
        cell.alignment = Alignment(horizontal=ALIGN_HORIZONTAL, vertical=ALIGN_VERTICAL)

def convert_to_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime.datetime):
        return value
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(str(value), fmt)
        except:
            continue
    return None

def get_last_day_previous_month(reference_date):
    first_day_this_month = reference_date.replace(day=1)
    return first_day_this_month - datetime.timedelta(days=1)

def calculate_review_status(last_login, reference_date):
    if not last_login:
        return "삭제", COLOR_RED
    delta_days = (reference_date - last_login).days
    if last_login > reference_date:
        return "유지", None
    if delta_days >= DAYS_DELETE:
        return "삭제", COLOR_RED
    elif delta_days >= DAYS_STOP:
        return "중지", COLOR_BLUE
    else:
        return "유지", None

def copy_data(source_sheet, target_sheet):
    headers = [cell.value if cell.value else "" for cell in source_sheet[1]]

    # source 컬럼 인덱스
    src_indices = {}
    for src_col in COLUMN_MAPPING.keys():
        try:
            src_indices[src_col] = headers.index(src_col)
        except ValueError:
            debug_log(f"Source column '{src_col}' not found. Skipping sheet.")
            return

    # target 컬럼 인덱스
    tgt_indices = {tgt_col: TARGET_COLUMNS.index(tgt_col)+1 for tgt_col in COLUMN_MAPPING.values()}

    today = datetime.datetime.today()
    reference_date = get_last_day_previous_month(today)

    for row_idx, row in enumerate(source_sheet.iter_rows(min_row=2, values_only=True), start=2):
        dt_val = None
        for src_col, tgt_col in COLUMN_MAPPING.items():
            src_idx = src_indices[src_col]
            tgt_idx = tgt_indices[tgt_col]
            value = str(row[src_idx]).strip() if row[src_idx] else ""

            if tgt_col == "마지막 로그인시간":
                dt_val = convert_to_datetime(value)
                value = dt_val.strftime("%Y-%m-%d %H:%M") if dt_val else ""

            cell = target_sheet.cell(row=row_idx, column=tgt_idx, value=value)
            cell.font = Font(size=FONT_SIZE)
            cell.alignment = Alignment(horizontal=ALIGN_HORIZONTAL, vertical=ALIGN_VERTICAL)

        # 검토결과
        review_idx = TARGET_COLUMNS.index("검토결과")+1
        review_val, color = calculate_review_status(dt_val, reference_date)
        review_cell = target_sheet.cell(row=row_idx, column=review_idx, value=review_val)
        review_cell.font = Font(size=FONT_SIZE)
        review_cell.alignment = Alignment(horizontal=ALIGN_HORIZONTAL, vertical=ALIGN_VERTICAL)
        if color:
            review_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        debug_log(f"Row {row_idx} processed: last_login={dt_val}, review={review_val}")

def process_excels():
    target_path = Path(TARGET_FILE)
    source_path = Path(SOURCE_FOLDER)

    if not target_path.exists():
        debug_log(f"Target file '{TARGET_FILE}' not found. Creating new workbook.")
        workbook_target = Workbook()
    else:
        debug_log(f"Loading target file: {TARGET_FILE}")
        workbook_target = load_workbook(target_path)

    sheet_target = ensure_worksheet(workbook_target, TARGET_SHEET)
    add_columns(sheet_target, TARGET_COLUMNS)

    for file_path in source_path.glob("*.xlsx"):
        debug_log(f"Processing source file: {file_path}")
        workbook_source = load_workbook(file_path, data_only=True)
        if SOURCE_SHEET not in workbook_source.sheetnames:
            debug_log(f"Sheet '{SOURCE_SHEET}' not found in {file_path}. Skipping...")
            continue
        sheet_source = workbook_source[SOURCE_SHEET]
        copy_data(sheet_source, sheet_target)

    workbook_target.save(target_path)
    debug_log(f"Saved target file: {TARGET_FILE}")

if __name__ == "__main__":
    process_excels()
