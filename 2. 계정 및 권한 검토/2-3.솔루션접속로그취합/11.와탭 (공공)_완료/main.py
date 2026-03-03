import os
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from config import (
    BASE_EXCEL_PATH,
    SOURCE_FOLDER_PATH,
    TARGET_SHEET_NAME,
    SOURCE_SHEET_INCLUDE,
    SOURCE_SHEET_EXCLUDES,
    TARGET_COLUMNS,
    SOURCE_COLUMN_MAPPING,
    FONT_SIZE,
    ALIGN_HORIZONTAL,
    ALIGN_VERTICAL,
    COLOR_RED,
    COLOR_BLUE,
    STOP_DAYS,
    DELETE_DAYS,
    OUTPUT_DATE_FORMAT
)


def log_debug(message):
    print(f"[DEBUG] {message}")


def normalize_value(value):
    if value is None:
        return ""
    return str(value).replace(" ", "").strip()


def get_last_day_of_previous_month():
    first_day = datetime.today().replace(day=1)
    return first_day - timedelta(days=1)


def extract_parenthesis_value(text):
    match = re.search(r"\((.*?)\)", text)
    return match.group(1) if match else ""


def ensure_target_sheet(workbook):
    if TARGET_SHEET_NAME not in workbook.sheetnames:
        log_debug("create target sheet")
        workbook.create_sheet(TARGET_SHEET_NAME)
    return workbook[TARGET_SHEET_NAME]


def apply_cell_style(cell):
    cell.font = Font(size=FONT_SIZE)
    cell.alignment = Alignment(
        horizontal=ALIGN_HORIZONTAL,
        vertical=ALIGN_VERTICAL
    )


def ensure_target_columns(sheet):
    headers = [cell.value for cell in sheet[1] if cell.value]
    col_index = len(headers)

    for col in TARGET_COLUMNS:
        if col not in headers:
            col_index += 1
            cell = sheet.cell(row=1, column=col_index)
            cell.value = col
            apply_cell_style(cell)
            headers.append(col)

    log_debug("columns ensured")


def parse_datetime(value):
    if not value:
        return None

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(str(value), fmt)
        except ValueError:
            continue
    return None


def evaluate_review_status(login_dt, base_date):
    if login_dt is None:
        return "삭제"

    diff_days = (base_date - login_dt).days

    if diff_days >= DELETE_DAYS:
        return "삭제"
    if diff_days >= STOP_DAYS:
        return "중지"
    return "유지"


def is_excluded_sheet(sheet_name):
    for exclude in SOURCE_SHEET_EXCLUDES:
        if exclude in sheet_name:
            return True
    return False


def process_source_files(target_sheet):
    base_date = get_last_day_of_previous_month()
    target_row = target_sheet.max_row + 1

    for file_name in os.listdir(SOURCE_FOLDER_PATH):
        if not file_name.endswith(".xlsx"):
            continue

        log_debug(f"open file: {file_name}")
        wb = load_workbook(os.path.join(SOURCE_FOLDER_PATH, file_name), data_only=True)

        for sheet_name in wb.sheetnames:
            if SOURCE_SHEET_INCLUDE in sheet_name and not is_excluded_sheet(sheet_name):
                log_debug(f"process sheet: {sheet_name}")
                sheet = wb[sheet_name]

                category_value = extract_parenthesis_value(sheet_name)
                headers = [cell.value for cell in sheet[1]]

                source_indexes = {
                    target: headers.index(source)
                    for source, target in SOURCE_COLUMN_MAPPING.items()
                    if source in headers
                }

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    login_dt = parse_datetime(
                        row[source_indexes["마지막 로그인시간"]]
                    )

                    review_result = evaluate_review_status(login_dt, base_date)

                    value_map = {
                        "ID": normalize_value(row[source_indexes["ID"]]),
                        "성명": normalize_value(row[source_indexes["성명"]]),
                        "구분": category_value,
                        "마지막 로그인시간": (
                            login_dt.strftime(OUTPUT_DATE_FORMAT)
                            if login_dt else ""
                        ),
                        "검토결과": review_result
                    }

                    for col_name, value in value_map.items():
                        col_idx = TARGET_COLUMNS.index(col_name) + 1
                        cell = target_sheet.cell(row=target_row, column=col_idx)
                        cell.value = value
                        apply_cell_style(cell)

                        if col_name == "검토결과":
                            if value == "중지":
                                cell.fill = PatternFill(
                                    fill_type="solid",
                                    start_color=COLOR_BLUE
                                )
                            elif value == "삭제":
                                cell.fill = PatternFill(
                                    fill_type="solid",
                                    start_color=COLOR_RED
                                )

                    target_row += 1


def main():
    log_debug("process start")

    workbook = load_workbook(BASE_EXCEL_PATH)
    target_sheet = ensure_target_sheet(workbook)

    ensure_target_columns(target_sheet)
    process_source_files(target_sheet)

    workbook.save(BASE_EXCEL_PATH)
    log_debug("process end")


if __name__ == "__main__":
    main()
