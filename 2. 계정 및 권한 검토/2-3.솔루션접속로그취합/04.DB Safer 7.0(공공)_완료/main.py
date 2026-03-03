import os
import logging
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from config import *

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s [DEBUG] %(message)s")


def normalize(value):
    if value is None:
        return ""
    return str(value).replace(" ", "").strip()


def get_last_day_previous_month():
    today = datetime.today()
    first_day_this_month = today.replace(day=1)
    return first_day_this_month - timedelta(days=1)


def parse_datetime(value):
    if pd.isna(value):
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(str(value), fmt)
        except:
            continue
    return None


def evaluate_review(dt, 기준일):
    if dt is None:
        return REVIEW_DELETE
    diff_days = (기준일 - dt).days
    if diff_days >= REVIEW_DELETE_DAYS:
        return REVIEW_DELETE
    if diff_days >= REVIEW_STOP_DAYS:
        return REVIEW_STOP
    return REVIEW_KEEP


def apply_style(cell, review_value=None):
    cell.font = Font(size=FONT_SIZE)
    cell.alignment = Alignment(horizontal=ALIGN_HORIZONTAL, vertical=ALIGN_VERTICAL)

    if review_value == REVIEW_STOP:
        cell.fill = PatternFill(start_color=COLOR_STOP, end_color=COLOR_STOP, fill_type="solid")
    elif review_value == REVIEW_DELETE:
        cell.fill = PatternFill(start_color=COLOR_DELETE, end_color=COLOR_DELETE, fill_type="solid")


def ensure_target_sheet_and_columns(wb):
    logging.debug("Ensuring target sheet and columns")
    if TARGET_SHEET_NAME not in wb.sheetnames:
        wb.create_sheet(TARGET_SHEET_NAME)
    ws = wb[TARGET_SHEET_NAME]

    for idx, col_name in enumerate(TARGET_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = col_name
        cell.font = Font(size=FONT_SIZE)
        cell.alignment = Alignment(horizontal=ALIGN_HORIZONTAL, vertical=ALIGN_VERTICAL)


def process_csv_files(wb):
    logging.debug("Start processing CSV files")
    ws_target = wb[TARGET_SHEET_NAME]
    기준일 = get_last_day_previous_month()
    target_row = ws_target.max_row + 1

    for file_name in os.listdir(SOURCE_CSV_DIR):
        if not file_name.endswith(".csv"):
            continue

        csv_path = os.path.join(SOURCE_CSV_DIR, file_name)
        df = pd.read_csv(csv_path, encoding="utf-8")

        logging.debug(f"Processing CSV: {file_name}, rows: {len(df)}")
        for _, row in df.iterrows():
            # ID
            val_id = normalize(row.get("[보안계정]"))
            ws_target.cell(row=target_row, column=TARGET_COLUMNS.index("ID") + 1, value=val_id)

            # 성명
            val_name = normalize(row.get("[사용자명]"))
            ws_target.cell(row=target_row, column=TARGET_COLUMNS.index("성명") + 1, value=val_name)

            # 마지막 로그인시간
            raw_dt = row.get("[마지막 로그인 시각]")
            parsed_dt = parse_datetime(raw_dt)
            formatted_dt = parsed_dt.strftime(OUTPUT_DATE_FORMAT) if parsed_dt else ""
            ws_target.cell(
                row=target_row,
                column=TARGET_COLUMNS.index("마지막 로그인시간") + 1,
                value=formatted_dt,
            )

            # 검토결과
            review = evaluate_review(parsed_dt, 기준일)
            cell_review = ws_target.cell(
                row=target_row,
                column=TARGET_COLUMNS.index("검토결과") + 1,
                value=review
            )

            # Apply style for all columns
            for col in range(1, len(TARGET_COLUMNS) + 1):
                apply_style(ws_target.cell(row=target_row, column=col),
                            review_value=review if TARGET_COLUMNS[col-1]=="검토결과" else None)

            target_row += 1


def main():
    logging.debug("Process start")
    wb = load_workbook(TARGET_EXCEL_FILE)
    ensure_target_sheet_and_columns(wb)
    process_csv_files(wb)
    wb.save(TARGET_EXCEL_FILE)
    logging.debug("Process end")


if __name__ == "__main__":
    main()
