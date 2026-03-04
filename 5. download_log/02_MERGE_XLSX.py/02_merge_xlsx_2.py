import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from copy import copy

# 📂 엑셀 파일 경로
folder_path = "./"  # ← 실제 폴더 경로
output_file = "merged_final.xlsx"

header_row = 3
max_columns = 44

# 컬럼명 설정 (실제 엑셀 컬럼명과 일치)
name_col = "이름"
email_col = "이메일 주소"
approval_col = "승인번호"  # 삭제 기준 컬럼

# 1️⃣ openpyxl로 데이터 + 서식 복사
merged_wb = Workbook()
merged_ws = merged_wb.active
merged_ws.title = "MergedData"
current_row = 1
headers = None

for file in os.listdir(folder_path):
    if file.endswith(".xlsx"):
        file_path = os.path.join(folder_path, file)
        print(f"📄 파일 읽는 중: {file_path}")

        wb = load_workbook(file_path)
        ws = wb.active

        # 3번째 행을 컬럼명으로 읽기
        local_headers = [ws.cell(row=header_row, column=i).value for i in range(1, max_columns + 1)]
        if headers is None:
            headers = local_headers
            # 헤더 복사
            for col, val in enumerate(headers, start=1):
                src_cell = ws.cell(row=header_row, column=col)
                dest_cell = merged_ws.cell(row=current_row, column=col, value=val)
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.number_format = copy(src_cell.number_format)
                    dest_cell.alignment = copy(src_cell.alignment)
            current_row += 1

        # 데이터 복사
        for r in range(header_row + 1, ws.max_row + 1):
            if all(ws.cell(row=r, column=c).value in [None, "", " "] for c in range(1, max_columns + 1)):
                continue
            for c in range(1, max_columns + 1):
                src_cell = ws.cell(row=r, column=c)
                dest_cell = merged_ws.cell(row=current_row, column=c, value=src_cell.value)
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.number_format = copy(src_cell.number_format)
                    dest_cell.alignment = copy(src_cell.alignment)
            current_row += 1

# 임시 저장
temp_file = os.path.join(folder_path, "temp_merged.xlsx")
merged_wb.save(temp_file)
print(f"✅ 1단계: 서식 포함 임시 병합 완료 → {temp_file}")

# 2️⃣ pandas로 중복 제거 및 승인번호 없는 행 삭제
df = pd.read_excel(temp_file, header=0)
df.columns = df.columns.astype(str).str.strip()  # 공백 제거

# 승인번호 컬럼 없는 행 삭제
if approval_col in df.columns:
    df = df[df[approval_col].notna() & (df[approval_col] != "")]
else:
    print(f"⚠️ '{approval_col}' 컬럼을 찾을 수 없습니다. 삭제하지 않습니다.")

# 이름+이메일 기준 중복 제거 (최근 행 유지)
if name_col in df.columns and email_col in df.columns:
    df.drop_duplicates(subset=[name_col, email_col], keep="last", inplace=True)

# 3️⃣ openpyxl로 삭제 반영 (서식 유지)
final_wb = load_workbook(temp_file)
final_ws = final_wb.active

# 삭제할 행 계산
valid_rows = set(df.index + 2)  # header 이후 실제 엑셀 행번호
rows_to_delete = [r for r in range(2, final_ws.max_row + 1) if r not in valid_rows]

# 뒤에서부터 삭제
for r in reversed(rows_to_delete):
    final_ws.delete_rows(r, 1)

# 최종 저장
output_path = os.path.join(folder_path, output_file)
final_wb.save(output_path)

print(f"\n🎉 승인번호 기준 필터 + 이름/이메일 중복 제거 완료!")
print(f"📁 최종 결과 파일: {output_path}")
