import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from config import CSV_DIR, USER_XLSX, LOG_XLSX, USER_SHEET, LOG_SHEET, FONT_SIZE, H_ALIGN, V_ALIGN

def debug(msg):
    print(f"[DEBUG] {msg}")

def create_user_excel():
    debug("Creating user Excel file...")
    wb = Workbook()
    ws = wb.active
    ws.title = USER_SHEET
    columns = ["ID", "성명", "마지막 로그인시간"]
    for col_num, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = Font(size=FONT_SIZE)
        cell.alignment = Alignment(horizontal=H_ALIGN, vertical=V_ALIGN)
    wb.save(USER_XLSX)
    debug(f"User Excel created at {USER_XLSX}")

def populate_user_excel():
    debug("Populating user Excel from CSV files...")
    wb = load_workbook(USER_XLSX)
    ws = wb[USER_SHEET]
    row_index = 2

    for file in os.listdir(CSV_DIR):
        if "object_user" in file and file.endswith(".csv"):
            debug(f"Processing file: {file}")
            df = pd.read_csv(os.path.join(CSV_DIR, file), dtype=str)
            for _, row in df.iterrows():
                ws.cell(row=row_index, column=1, value=row["ID"].strip())
                ws.cell(row=row_index, column=2, value=row["USER NAME"].strip())
                row_index += 1

    wb.save(USER_XLSX)
    debug("User Excel populated.")

def create_log_excel():
    debug("Creating log Excel file...")
    wb = Workbook()
    ws = wb.active
    ws.title = LOG_SHEET
    wb.save(LOG_XLSX)
    debug(f"Log Excel created at {LOG_XLSX}")

def populate_log_excel():
    debug("Populating log Excel from CSV files...")
    combined_df = pd.DataFrame()
    for file in os.listdir(CSV_DIR):
        if "Ahnlab" in file and file.endswith(".csv"):
            debug(f"Processing log file: {file}")
            df = pd.read_csv(os.path.join(CSV_DIR, file), dtype=str)
            # 필터링: description에 "사용자가 로그인했습니다"
            df = df[df["description"].str.strip() == "사용자가 로그인했습니다"]
            combined_df = pd.concat([combined_df, df], ignore_index=True)

    if combined_df.empty:
        debug("No log entries found.")
        return

    # timestamp를 datetime으로 변환, NaT 제거
    combined_df["timestamp"] = pd.to_datetime(combined_df["timestamp"], errors="coerce")
    combined_df = combined_df[combined_df["timestamp"].notna()]

    # user_id 기준 최신 timestamp만 남김
    combined_df.sort_values(by=["user_id", "timestamp"], ascending=[True, False], inplace=True)
    combined_df = combined_df.drop_duplicates(subset="user_id", keep="first")

    # 엑셀 저장
    combined_df.to_excel(LOG_XLSX, sheet_name=LOG_SHEET, index=False)
    debug("Log Excel populated.")

def map_timestamp_to_user():
    debug("Mapping timestamp to user Excel...")
    wb_user = load_workbook(USER_XLSX)
    ws_user = wb_user[USER_SHEET]

    df_log = pd.read_excel(LOG_XLSX, sheet_name=LOG_SHEET, dtype=str)
    df_log["timestamp"] = pd.to_datetime(df_log["timestamp"], errors="coerce")

    for row in range(2, ws_user.max_row + 1):
        user_id = str(ws_user.cell(row=row, column=1).value).strip()
        log_row = df_log[df_log["user_id"].str.strip() == user_id]
        if not log_row.empty:
            ts = log_row.iloc[0]["timestamp"]
            if pd.notna(ts):
                ws_user.cell(row=row, column=3, value=ts.strftime("%Y-%m-%d %H:%M"))

    wb_user.save(USER_XLSX)
    debug("Timestamp mapping completed.")

def main():
    create_user_excel()
    populate_user_excel()
    create_log_excel()
    populate_log_excel()
    map_timestamp_to_user()

if __name__ == "__main__":
    main()
