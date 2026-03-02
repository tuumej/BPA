# @@작성자 : tuumej
# @@작성일자 : 260113
# @@참고_01 : 스크립트
# 방화벽 정책 요청(ITSM) 정렬 및 VPC 정보 출력

# @@참고_02 : exe 파일 만드는 방법
# exe 파일 생성을 위한 사전 작업 : pyinstaller --onefile main.py
# exe 파일 생성 : pyinstaller --onefile main.py
import pandas as pd
import ipaddress
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill


# =========================
# VPC 정보
# =========================
vpc_info = {
    # 검토 IP대역
    "xxx-vpc-01"   : "xxx.xxx.xxx.xxx/21",
    "xxx-vpc-02"  : "xxx.xxx.xxx.xxx/24",
}


# =========================
# 경로 설정
# =========================
BASE_DIR = Path.cwd()
DATA_DIR = BASE_DIR / "data"
OUTPUT_EXCEL_PATH = BASE_DIR / "result.xlsx"

INPUT_EXCEL_FILES = (
    list(DATA_DIR.glob("*.xlsx")) +
    list(DATA_DIR.glob("*.xls"))
)


# =========================
# 유틸 함수
# =========================
def find_vpc_name(ip_str):
    if pd.isna(ip_str):
        return None
    try:
        ip = ipaddress.ip_address(str(ip_str))
    except ValueError:
        return None

    for name, cidr in vpc_info.items():
        if ip in ipaddress.ip_network(cidr):
            return name
    return None


def judge_result(vpc_a, vpc_d):
    if pd.notna(vpc_a) and pd.notna(vpc_d) and vpc_a == vpc_d:
        return "ACG"
    return "SFC + ACG"


# =========================
# 메인 처리 로직
# =========================
def process_excel(input_files, output_path):
    dfs = []

    # 1. 모든 엑셀 파일 읽기
    for file in input_files:
        df = pd.read_excel(file)
        df = df[["A", "B", "D", "E", "G"]]
        dfs.append(df)

    if not dfs:
        raise ValueError("data 폴더에 엑셀 파일이 없습니다.")

    # 2. 병합
    df = pd.concat(dfs, ignore_index=True)

    # 3. D 기준 정렬
    df = df.sort_values(by="D").reset_index(drop=True)

    # 4. VPC 매핑
    df["vpc_info_A"] = df["A"].apply(find_vpc_name)
    df["vpc_info_D"] = df["D"].apply(find_vpc_name)

    # 5. 결과 컬럼 추가 (맨 우측)
    df["결과"] = df.apply(
        lambda r: judge_result(r["vpc_info_A"], r["vpc_info_D"]),
        axis=1
    )

    # 컬럼 위치 조정 (A, D 오른쪽)
    cols = list(df.columns)
    cols.insert(cols.index("A") + 1, cols.pop(cols.index("vpc_info_A")))
    cols.insert(cols.index("D") + 1, cols.pop(cols.index("vpc_info_D")))
    df = df[cols]

    # 6. 완전 중복 행 하단 이동
    dup_mask = df.duplicated(
        subset=["A", "B", "D", "E", "G"],
        keep=False
    )

    df_unique = df[~dup_mask]
    df_dup = df[dup_mask]

    rows = []
    for _, row in df_dup.iterrows():
        rows.append(row)
        rows.append(pd.Series([None] * len(df.columns), index=df.columns))

    df_dup_spaced = pd.DataFrame(rows)

    df_result = pd.concat(
        [df_unique, pd.DataFrame([{}]), df_dup_spaced],
        ignore_index=True
    )

    # 7. 엑셀 저장
    df_result.to_excel(output_path, index=False)

    # =========================
    # 엑셀 서식 + 색칠
    # =========================
    wb = load_workbook(output_path)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    # 열 너비 15
    for col in range(1, max_col + 1):
        ws.column_dimensions[
            ws.cell(row=1, column=col).column_letter
        ].width = 15

    # 테두리
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 색상
    yellow = PatternFill("solid", fgColor="FFFF99")
    blue = PatternFill("solid", fgColor="BDD7EE")
    fills = [yellow, blue]

    prev_key = None
    color_idx = 0

    for row in range(2, max_row + 1):
        d = ws[f"D{row}"].value
        e = ws[f"E{row}"].value
        g = ws[f"G{row}"].value

        if d is None and e is None and g is None:
            continue

        key = (d, e, g)

        if prev_key is None:
            color_idx = 0
        elif key != prev_key:
            color_idx = 1 - color_idx

        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fills[color_idx]
            cell.border = thin_border

        prev_key = key

    # 헤더 테두리
    for col in range(1, max_col + 1):
        ws.cell(row=1, column=col).border = thin_border

    wb.save(output_path)


# =========================
# main
# =========================
def main():
    try:
        print("엑셀 처리 시작")
        process_excel(INPUT_EXCEL_FILES, OUTPUT_EXCEL_PATH)
        print(f"처리 완료 → {OUTPUT_EXCEL_PATH}")
    except Exception as e:
        print(f"처리 실패: {e}")


if __name__ == "__main__":
    main()
