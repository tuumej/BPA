import pandas as pd
from openpyxl import load_workbook

from config import (
    EXCEL_FILE_PATH,
    TARGET_SHEET_NAME,
    NAME_COLUMN,
    ID_COLUMN
)


def main():
    try:
        # 엑셀 파일 로드
        excel_file = pd.ExcelFile(EXCEL_FILE_PATH)

        collected_rows = []

        # 모든 시트 순회
        for sheet_name in excel_file.sheet_names:
            if sheet_name == TARGET_SHEET_NAME:
                continue

            df = excel_file.parse(sheet_name=sheet_name)

            # 필수 컬럼 존재 여부 확인
            if NAME_COLUMN in df.columns and ID_COLUMN in df.columns:
                collected_rows.append(
                    df[[NAME_COLUMN, ID_COLUMN]]
                )

        if not collected_rows:
            print("수집할 데이터가 없습니다.")
            return

        # 데이터 병합
        merged_df = pd.concat(collected_rows, ignore_index=True)

        # ID 기준 중복 제거
        merged_df = merged_df.drop_duplicates(subset=[ID_COLUMN])

        # 기존 엑셀 열기
        workbook = load_workbook(EXCEL_FILE_PATH)

        # 기존 시트 삭제 (있다면)
        if TARGET_SHEET_NAME in workbook.sheetnames:
            del workbook[TARGET_SHEET_NAME]
            workbook.save(EXCEL_FILE_PATH)

        # 새 시트로 저장
        with pd.ExcelWriter(
            EXCEL_FILE_PATH,
            engine="openpyxl",
            mode="a"
        ) as writer:
            merged_df.to_excel(
                writer,
                sheet_name=TARGET_SHEET_NAME,
                index=False
            )

        print("00.인력관리 시트 생성 및 중복 제거 완료")

    except Exception as e:
        print(f"[ERROR] {e}")


if __name__ == "__main__":
    main()
