import os
import json
import gzip
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from config import (
    BASE_SEARCH_DIR,
    MERGED_JSON_PATH,
    RESULT_XLSX_PATH,
    RESULT_SHEET_NAME,
    FONT_SIZE,
    ALIGN_HORIZONTAL,
    ALIGN_VERTICAL,
    DEBUG_ENABLE,
    ESCAPED_QUOTE_PATTERN,
    REPLACED_QUOTE,
    MESSAGE_KEY,
    TARGET_DATETIME_FORMAT
)

# =========================
# Debug
# =========================
def debug(message):
    if DEBUG_ENABLE:
        print(f"[DEBUG] {message}")

# =========================
# Normalize
# =========================
def normalize_string(value):
    return value.replace(ESCAPED_QUOTE_PATTERN, REPLACED_QUOTE).strip()

# =========================
# Datetime Parse
# =========================
def parse_datetime(value):
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y.%m.%d %H:%M:%S",
        "%Y.%m.%d %H:%M",
        "%Y.%m.%d"
    ]

    for fmt in formats:
        try:
            return datetime.strptime(value, fmt).strftime(TARGET_DATETIME_FORMAT)
        except ValueError:
            continue

    return value

# =========================
# KV Parser
# =========================
def parse_kv_pairs(value):
    pairs = []
    for item in value.split(","):
        item = item.strip()
        if ":" not in item:
            continue
        key, val = item.split(":", 1)
        pairs.append((key.strip(), val.strip()))
    return pairs

# =========================
# JSON Flatten
# =========================
def flatten_json(data, parent_key=""):
    items = {}

    if isinstance(data, dict):
        for key, value in data.items():
            new_key = f"{parent_key}.{key}" if parent_key else key

            if isinstance(value, dict) or isinstance(value, list):
                items.update(flatten_json(value, new_key))
                continue

            if isinstance(value, str):
                value = normalize_string(value)

                if key == MESSAGE_KEY and "," in value and ":" in value:
                    pairs = parse_kv_pairs(value)

                    if pairs:
                        first_key, first_val = pairs[0]
                        items[new_key] = parse_datetime(first_val)

                        for sub_key, sub_val in pairs[1:]:
                            items[f"{new_key}.{sub_key}"] = sub_val
                    else:
                        items[new_key] = value

                    continue

                if "," in value and ":" in value:
                    for sub_key, sub_val in parse_kv_pairs(value):
                        items[f"{new_key}.{sub_key}"] = sub_val
                else:
                    items[new_key] = value
            else:
                items[new_key] = value

    elif isinstance(data, list):
        for index, value in enumerate(data):
            items.update(flatten_json(value, f"{parent_key}.{index}"))

    return items

# =========================
# GZ Collector
# =========================
def collect_gz_files():
    debug("search gz files")
    gz_files = []

    for root, dirs, files in os.walk(BASE_SEARCH_DIR):
        if "" not in root:
            continue
        for file in files:
            if file.endswith(".gz"):
                gz_files.append(os.path.join(root, file))

    debug(f"found gz files: {len(gz_files)}")
    return gz_files

# =========================
# Merge GZ
# =========================
def merge_gz_files(gz_files):
    debug("merge gz files")
    merged = []

    for path in gz_files:
        debug(f"read gz: {path}")
        with gzip.open(path, "rt", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                obj = json.loads(line)
                merged.append(flatten_json(obj))

    with open(MERGED_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)

    debug("merged json created")
    return merged

# =========================
# Write Excel
# =========================
def write_excel(data):
    debug("write excel start")

    wb = Workbook()
    ws = wb.active
    ws.title = RESULT_SHEET_NAME

    headers = []
    for row in data:
        for key in row.keys():
            if key not in headers:
                headers.append(key)

    font = Font(size=FONT_SIZE)
    align = Alignment(horizontal=ALIGN_HORIZONTAL, vertical=ALIGN_VERTICAL)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = font
        cell.alignment = align

    for row_idx, row in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
            cell.font = font
            cell.alignment = align

    wb.save(RESULT_XLSX_PATH)
    debug("write excel end")

# =========================
# Main
# =========================
def main():
    debug("program start")

    gz_files = collect_gz_files()
    if not gz_files:
        debug("no gz files")
        return

    merged_data = merge_gz_files(gz_files)
    if not merged_data:
        debug("no merged data")
        return

    write_excel(merged_data)
    debug("program end")

if __name__ == "__main__":
    main()
