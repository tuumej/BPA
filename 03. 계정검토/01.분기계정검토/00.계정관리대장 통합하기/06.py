import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side



INPUT_DIR = r""     # 엑셀 파일들이 있는 폴더
OUTPUT_FILE = r""

EMAIL_COLUMN = "이메일 주소"

# ✅ 결과에 포함할 컬럼 (병합 대상)
MERGE_COLUMNS = [
    "승인번호"
]

TOP_ROW_VALUES = [
    "-"
]


# =========================
# 유틸 함수
# =========================
def merge_values(series: pd.Series) -> str:
    values = (
        series.dropna()
        .astype(str)
        .map(str.strip)
        .replace("-", "")
    )
    values = [v for v in values if v]

    if not values:
        return "-"

    unique = sorted(set(values))
    return unique[0] if len(unique) == 1 else ",".join(unique)


def make_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols, seen = [], {}
    for col in df.columns:
        if col not in seen:
            seen[col] = 0
            cols.append(col)
        else:
            seen[col] += 1
            cols.append(f"{col}_{seen[col]}")
    df.columns = cols
    return df


def read_excel_with_header_3row(filepath: str) -> pd.DataFrame:
    df = pd.read_excel(filepath, header=2)
    df = make_unique_columns(df)

    # 🔥 MERGE_COLUMNS만 유지
    return df[[c for c in MERGE_COLUMNS if c in df.columns]]


# =========================
# 메인 로직
# =========================
def main():
    dfs = []

    for filename in os.listdir(INPUT_DIR):
        if filename.lower().endswith((".xlsx", ".xls")):
            try:
                df = read_excel_with_header_3row(os.path.join(INPUT_DIR, filename))
                dfs.append(df)
            except Exception as e:
                print(f"❌ 실패: {filename} / {e}")

    if not dfs:
        print("❌ 병합할 엑셀 없음")
        return

    merged_df = pd.concat(dfs, ignore_index=True)

    # 이메일 정제
    merged_df[EMAIL_COLUMN] = merged_df[EMAIL_COLUMN].astype(str).str.strip()

    # =========================
    # 이메일 기준 병합 제외 조건
    # =========================
    invalid_mask = (
        merged_df[EMAIL_COLUMN].isna() |
        (merged_df[EMAIL_COLUMN] == "") |
        (merged_df[EMAIL_COLUMN] == "-")
    )

    valid_df = merged_df[~invalid_mask]
    invalid_df = merged_df[invalid_mask]

    agg_rules = {
        col: merge_values
        for col in MERGE_COLUMNS
        if col != EMAIL_COLUMN and col in valid_df.columns
    }

    merged_valid_df = (
        valid_df
        .groupby(EMAIL_COLUMN)
        .agg(agg_rules)
        .reset_index()
    )

    final_df = pd.concat([merged_valid_df, invalid_df], ignore_index=True)

    # =========================
    # NO 컬럼 추가 + 컬럼 순서 고정
    # =========================
    final_df.insert(0, "NO", range(1, len(final_df) + 1))
    final_df = final_df[["NO"] + MERGE_COLUMNS]

    final_df = final_df.fillna("-")
    final_df.to_excel(OUTPUT_FILE, index=False)

    # =========================
    # 엑셀 서식 처리
    # =========================
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    ws.insert_rows(1)
    for i, v in enumerate(TOP_ROW_VALUES, start=1):
        ws.cell(row=1, column=i, value=v)

    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in ws.iter_rows():
        for cell in row:
            if cell.value in [None, ""]:
                cell.value = "-"
            cell.alignment = align
            cell.border = border

    wb.save(OUTPUT_FILE)
    print(f"✅ 최종 엑셀 생성 완료: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
