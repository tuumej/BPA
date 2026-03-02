import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from config import (
    TARGET_EXCEL_PATH,
    RESULT_FOLDER_PATH,
    TARGET_WORKSHEET_NAME,
    CREATE_COLUMNS,
    RESULT_SOURCE_COLUMNS,
    FONT_SIZE,
    ALIGN_HORIZONTAL,
    ALIGN_VERTICAL,
    COLOR_RED,
    COLOR_BLUE,
    RESULT_KEEP,
    RESULT_STOP,
    RESULT_DELETE,
    DATETIME_FORMAT
)

def log(message):
    print(f"[DEBUG] {message}")

def get_or_create_worksheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        log(f"Worksheet exists: {sheet_name}")
        return wb[sheet_name]
    log(f"Worksheet created: {sheet_name}")
    return wb.create_sheet(sheet_name)

def apply_cell_style(cell):
    cell.font = Font(size=FONT_SIZE)
    cell.alignment = Alignment(
        horizontal=ALIGN_HORIZONTAL,
        vertical=ALIGN_VERTICAL
    )

def create_columns(ws):
    log("Creating columns")

    headers = []
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).strip() != "":
            headers.append(cell.value)

    for col_name in CREATE_COLUMNS:
        if col_name not in headers:
            col_idx = len(headers) + 1
            ws.cell(row=1, column=col_idx, value=col_name)
            apply_cell_style(ws.cell(row=1, column=col_idx))
            headers.append(col_name)
            log(f"Column added: {col_name}")

def get_last_day_previous_month():
    today = datetime.today()
    first_day_this_month = today.replace(day=1)
    return first_day_this_month - timedelta(days=1)

def parse_datetime(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    value = str(value).strip()
    if value == "":
        return None

    formats = [
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y.%m.%d %H:%M",
        "%Y.%m.%d"
    ]

    for fmt in formats:
        try:
            return datetime.strptime(value, fmt)
        except Exception:
            continue

    log(f"Datetime parse failed: {value}")
    return None

def evaluate_result(login_time, 기준일):
    if login_time is None:
        return RESULT_DELETE

    if login_time > 기준일:
        return RESULT_KEEP

    delta_days = (기준일 - login_time).days

    if delta_days >= 180:
        return RESULT_DELETE
    if delta_days >= 90:
        return RESULT_STOP
    return RESULT_KEEP

def load_result_files():
    files = []
    for file in os.listdir(RESULT_FOLDER_PATH):
        if "Result" in file and file.endswith(".xlsx"):
            files.append(os.path.join(RESULT_FOLDER_PATH, file))
    log(f"Result files found: {len(files)}")
    return files

def extract_result_data():
    result_data = []
    files = load_result_files()

    for file_path in files:
        log(f"Reading file: {file_path}")
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        name_idx = headers.index("이름")
        id_idx = headers.index("아이디")
        date_idx = headers.index("최종접속일자")

        for row in ws.iter_rows(min_row=2, values_only=True):
            result_data.append({
                "성명": str(row[name_idx]).strip() if row[name_idx] else "",
                "ID": str(row[id_idx]).strip() if row[id_idx] else "",
                "마지막 로그인시간": row[date_idx]
            })

    log(f"Extracted rows: {len(result_data)}")
    return result_data

def write_target_data(ws, data):
    headers = [cell.value for cell in ws[1]]

    name_col = headers.index("성명") + 1
    id_col = headers.index("ID") + 1
    login_col = headers.index("마지막 로그인시간") + 1
    result_col = headers.index("검토결과") + 1

    기준일 = get_last_day_previous_month()
    log(f"Base date: {기준일.strftime('%Y-%m-%d')}")

    start_row = ws.max_row + 1

    for idx, row_data in enumerate(data):
        row_num = start_row + idx

        ws.cell(row=row_num, column=name_col, value=row_data["성명"])
        ws.cell(row=row_num, column=id_col, value=row_data["ID"])

        login_time = parse_datetime(row_data["마지막 로그인시간"])
        if login_time:
            ws.cell(
                row=row_num,
                column=login_col,
                value=login_time.strftime(DATETIME_FORMAT)
            )

        result_value = evaluate_result(login_time, 기준일)
        result_cell = ws.cell(row=row_num, column=result_col, value=result_value)

        if result_value == RESULT_STOP:
            result_cell.fill = PatternFill(start_color=COLOR_BLUE, end_color=COLOR_BLUE, fill_type="solid")
        elif result_value == RESULT_DELETE:
            result_cell.fill = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid")

        for col in [name_col, id_col, login_col, result_col]:
            apply_cell_style(ws.cell(row=row_num, column=col))

    log("Target worksheet updated")

def main():
    log("Process started")

    wb = load_workbook(TARGET_EXCEL_PATH)
    ws = get_or_create_worksheet(wb, TARGET_WORKSHEET_NAME)

    create_columns(ws)
    result_data = extract_result_data()
    write_target_data(ws, result_data)

    wb.save(TARGET_EXCEL_PATH)
    log("Process completed")

if __name__ == "__main__":
    main()
