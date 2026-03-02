import os
import re
import logging
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from config import *

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s [DEBUG] %(message)s")


def normalize(value):
    if value is None:
        return ""
    return str(value).replace(" ", "").strip()


def get_last_day_previous_month():
    today = datetime.today()
    first_day_this_month = today.replace(day=1)
    return first_day_this_month - timedelta(days=1)


def extract_bracket_value(text):
    match = re.search(r"\((.*?)\)", text)
    return match.group(1) if match else ""


def ensure_target_sheet_and_columns(wb):
    logging.debug("Ensure target sheet and columns")
    if TARGET_SHEET_NAME not in wb.sheetnames:
        wb.create_sheet(title=TARGET_SHEET_NAME)
    ws = wb[TARGET_SHEET_NAME]

    for idx, col_name in enumerate(TARGET_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = col_name
        cell.font = Font(size=FONT_SIZE)
        cell.alignment = Alignment(horizontal=ALIGN_HORIZONTAL, vertical=ALIGN_VERTICAL)


def parse_datetime(value):
    if not value:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(str(value), fmt)
        except:
            continue
    return None


def evaluate_review(dt, 기준일):
    if dt is None:
        return REVIEW_DELETE
    diff_days = (기준일 - dt).days
    if diff_days >= REVIEW_DELETE_DAYS:
        return REVIEW_DELETE
    if diff_days >= REVIEW_STOP_DAYS:
        return REVIEW_STOP
    return REVIEW_KEEP


def apply_style(cell, review_value=None):
    cell.font = Font(size=FONT_SIZE)
    cell.alignment = Alignment(horizontal=ALIGN_HORIZONTAL, vertical=ALIGN_VERTICAL)

    if review_value == REVIEW_STOP:
        cell.fill = PatternFill(start_color=COLOR_STOP, end_color=COLOR_STOP, fill_type="solid")
    elif review_value == REVIEW_DELETE:
        cell.fill = PatternFill(start_color=COLOR_DELETE, end_color=COLOR_DELETE, fill_type="solid")


def process_source_excels(wb):
    logging.debug("Start processing source excels")
    ws_target = wb[TARGET_SHEET_NAME]
    기준일 = get_last_day_previous_month()
    target_row = ws_target.max_row + 1

    for file_name in os.listdir(SOURCE_EXCEL_DIR):
        if not file_name.endswith(".xlsx"):
            continue

        src_path = os.path.join(SOURCE_EXCEL_DIR, file_name)
        src_wb = load_workbook(src_path, data_only=True)

        for sheet_name in src_wb.sheetnames:
            if sheet_name.startswith("##"):  # 건너뛸 주석 시트
                continue

            normalized_name = normalize(sheet_name)

            # include/exclude filter
            if not all(k.replace(" ", "") in normalized_name for k in INCLUDE_KEYWORDS):
                continue
            if any(k.replace(" ", "") in normalized_name for k in EXCLUDE_KEYWORDS):
                continue

            logging.debug(f"Processing sheet: {sheet_name}")
            ws_src = src_wb[sheet_name]
            headers = [normalize(c.value) for c in ws_src[1]]

            # get column indexes
            col_idx_map = {}
            for src_col, tgt_col in COLUMN_MAPPING.items():
                if src_col in headers:
                    col_idx_map[tgt_col] = headers.index(src_col)

            # copy rows
            for row in ws_src.iter_rows(min_row=2, values_only=True):
                # ID
                val_id = normalize(row[col_idx_map.get("ID", 0)])
                ws_target.cell(row=target_row, column=TARGET_COLUMNS.index("ID") + 1, value=val_id)

                # 성명
                val_name = normalize(row[col_idx_map.get("성명", 0)])
                ws_target.cell(row=target_row, column=TARGET_COLUMNS.index("성명") + 1, value=val_name)

                # 마지막 로그인시간
                raw_dt = row[col_idx_map.get("마지막 로그인시간", 0)]
                parsed_dt = parse_datetime(raw_dt)
                formatted_dt = parsed_dt.strftime(OUTPUT_DATE_FORMAT) if parsed_dt else ""
                ws_target.cell(
                    row=target_row, column=TARGET_COLUMNS.index("마지막 로그인시간") + 1, value=formatted_dt
                )

                # 검토결과
                review = evaluate_review(parsed_dt, 기준일)
                cell_review = ws_target.cell(row=target_row, column=TARGET_COLUMNS.index("검토결과") + 1, value=review)

                # 구분 (괄호 안 값)
                ws_target.cell(
                    row=target_row,
                    column=TARGET_COLUMNS.index("구분") + 1,
                    value=extract_bracket_value(sheet_name)
                )

                # apply style for all columns
                for col in range(1, len(TARGET_COLUMNS) + 1):
                    apply_style(ws_target.cell(row=target_row, column=col),
                                review_value=review if TARGET_COLUMNS[col-1]=="검토결과" else None)

                target_row += 1


def main():
    logging.debug("Process start")
    wb = load_workbook(TARGET_EXCEL_FILE)
    ensure_target_sheet_and_columns(wb)
    process_source_excels(wb)
    wb.save(TARGET_EXCEL_FILE)
    logging.debug("Process end")


if __name__ == "__main__":
    main()
