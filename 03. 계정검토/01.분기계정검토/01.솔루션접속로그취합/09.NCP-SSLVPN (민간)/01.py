from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from config import (
    TARGET_EXCEL_PATH,
    RESULT_EXCEL_PATH,
    TARGET_SHEET_NAME,
    RESULT_SHEET_NAME,
    TARGET_COLUMNS,
    RESULT_FILTER_COLUMN,
    RESULT_FILTER_VALUE,
    RESULT_MAPPING,
    FONT_SIZE,
    ALIGN_HORIZONTAL,
    ALIGN_VERTICAL,
    COLOR_RED,
    COLOR_BLUE
)

# ================= DEBUG =================
def print_debug(msg):
    print(f"[DEBUG] {msg}")

# ================= JSON FLATTEN =================
def flatten_json(data, parent_key="", result=None):
    if result is None:
        result = {}

    if isinstance(data, dict):
        for k, v in data.items():
            new_key = f"{parent_key}.{k}" if parent_key else k
            flatten_json(v, new_key, result)

    elif isinstance(data, list):
        for item in data:
            flatten_json(item, parent_key, result)

    else:
        value = str(data).replace('"', '').strip()
        result[parent_key] = value

    return result

# ================= DATE PARSE =================
def parse_datetime(value):
    if value is None:
        return None

    value = str(value).strip()
    if value == "":
        return None

    for fmt in [
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y.%m.%d",
        "%Y-%m-%d %H:%M:%S"
    ]:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None

# ================= MAIN =================
def main():
    print_debug("start")

    wb_target = load_workbook(TARGET_EXCEL_PATH)

    if TARGET_SHEET_NAME in wb_target.sheetnames:
        ws_target = wb_target[TARGET_SHEET_NAME]
    else:
        ws_target = wb_target.create_sheet(TARGET_SHEET_NAME)

    header_row = 1
    existing_headers = [cell.value for cell in ws_target[header_row] if cell.value]

    for col in TARGET_COLUMNS:
        if col not in existing_headers:
            ws_target.cell(
                row=header_row,
                column=len(existing_headers) + 1,
                value=col
            )
            existing_headers.append(col)

    header_index = {v: i + 1 for i, v in enumerate(existing_headers)}

    font_style = Font(size=FONT_SIZE)
    align_style = Alignment(
        horizontal=ALIGN_HORIZONTAL,
        vertical=ALIGN_VERTICAL
    )

    wb_result = load_workbook(RESULT_EXCEL_PATH)
    ws_result = wb_result[RESULT_SHEET_NAME]

    result_headers = [cell.value for cell in ws_result[1]]
    filter_idx = result_headers.index(RESULT_FILTER_COLUMN)

    today = datetime.today()
    base_date = (today.replace(day=1) - timedelta(days=1)).replace(
        hour=0, minute=0
    )

    target_row = ws_target.max_row + 1

    for row in ws_result.iter_rows(min_row=2, values_only=True):
        if str(row[filter_idx]).strip() != RESULT_FILTER_VALUE:
            continue

        for src, dst in RESULT_MAPPING.items():
            if src not in result_headers:
                continue

            src_idx = result_headers.index(src)
            ws_target.cell(
                row=target_row,
                column=header_index[dst],
                value=str(row[src_idx]).replace('"', '').strip()
            )

        login_cell = ws_target.cell(
            row=target_row,
            column=header_index["마지막 로그인시간"]
        )

        login_dt = parse_datetime(login_cell.value)

        review_cell = ws_target.cell(
            row=target_row,
            column=header_index["검토결과"]
        )

        if login_dt is None:
            review_cell.value = "삭제"
            review_cell.fill = PatternFill(
                start_color=COLOR_RED,
                fill_type="solid"
            )
        else:
            diff_days = (base_date - login_dt).days

            if login_dt > base_date:
                review_cell.value = "유지"

            elif diff_days >= 180:
                review_cell.value = "삭제"
                review_cell.fill = PatternFill(
                    start_color=COLOR_RED,
                    fill_type="solid"
                )

            elif diff_days >= 90:
                review_cell.value = "중지"
                review_cell.fill = PatternFill(
                    start_color=COLOR_BLUE,
                    fill_type="solid"
                )
            else:
                review_cell.value = "유지"

        for col in header_index.values():
            cell = ws_target.cell(row=target_row, column=col)
            cell.font = font_style
            cell.alignment = align_style

        target_row += 1

    wb_target.save(TARGET_EXCEL_PATH)
    print_debug("write excel file completed")

if __name__ == "__main__":
    main()
